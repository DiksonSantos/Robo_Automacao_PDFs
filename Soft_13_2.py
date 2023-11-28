import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from pdf2image import convert_from_path
import pytesseract
from PIL import Image
import cv2
import numpy as np
import os
import openpyxl
import time
from datetime import datetime
import sys
import datetime
import pickle
import re

from Valor_Comprovante_CX import extrair_valor
from Data_Comprov_BB import encontrar_data_apos_palavra_chave, chave

Hora_inicial = datetime.datetime.now()
Hora_inicial = Hora_inicial.strftime("%H:%M")
tempo_inicial = time.time()


# Função para encontrar a primeira linha vazia em uma coluna específica
def encontrar_primeira_linha_vazia(sheet, coluna):
    for linha in sheet.iter_rows(
        min_row=1, max_col=sheet.max_column, max_row=sheet.max_row
    ):
        for cell in linha:
            if cell.column_letter == coluna and cell.value is None:
                return cell.row
    return sheet.max_row + 1


final_2 = []
final_3 = []


ORGAO = []
OBJETO = []
CREDOR = []
VALOR_EMP = []
RECURSO = []
DATA_EMP = []
DATA_NF = []
DATA_PG = []
AG = []
#CONTA = []
VALOR_COMPR = []
N_EMPRENHO = []
#################################

################################

# Função para processar um arquivo PDF
def processar_pdf(pdf_path, sheet):
    """A extensão.PDF deve estar em Maiuscula"""
    # Verifique se o arquivo PDF especificado existe
    #final_2 = []

    # global final_3
    if not os.path.exists(pdf_path) or not pdf_path.endswith(".PDF"):
        print(f"Arquivo PDF não encontrado ou extensão incorreta: {pdf_path}")
        return final_2

    try:
        # ARQUIVO DE SAÍDA:
        output_file = "texto_extraido.txt"

        # Use pdf2image para converter todas as páginas do PDF em imagens
        pages_as_images = convert_from_path(pdf_path)

        # Crie um arquivo de texto para salvar o texto extraído
        with open(output_file, "w", encoding="utf-8") as text_file:
            # Itere sobre as páginas e extraia o texto de cada uma
            for page_number, page_image in enumerate(pages_as_images):
                # Converta a imagem em escala de cinza
                gray_image = cv2.cvtColor(np.array(page_image), cv2.COLOR_RGB2GRAY)

                # Use o PyTesseract para extrair texto da imagem em escala de cinza
                texto = pytesseract.image_to_string(Image.fromarray(gray_image))

                # Salve o texto extraído no arquivo de texto
                text_file.write(f"Texto da página {page_number + 1}:\n{texto}\n\n")

        # Inicialize a lista final_2
        # final_2 = []
        caminho_arquivo = "texto_extraido.txt"


        ############################################################################################

        try:
            # CAPTANDO VALOR DO EMPENHO   COL 'H'
            # Data do Contrato: R$ 1.439,51 **R$ 1.439,52** R$ 0,03 (Valor do Meio)
            time.sleep(1)
            valor_empenho = None
            with open(caminho_arquivo, "r") as arquivo:
                # Leia o arquivo linha por linha
                linhas = arquivo.readlines()

                for linha in linhas:
                    # Verifique se a linha contém a informação "Valor do Empenho"
                    if "Data do Contrato:" in linha:
                        linha = linha.replace("S", "$")
                        valor_empenho = (
                            linha.split("R$")[2].strip() if "R$" in linha else None
                        )
                        VALOR_EMP.insert(0, valor_empenho)
                        break

                if valor_empenho is None:
                    VALOR_EMP.insert(0, "R$ Val_Empenho_")
            # print('Final_2: ',final_2)

            # Valor do Empenho: R$ 125,00  (MESMA LINHA)
            time.sleep(1)
            with open(caminho_arquivo, "r") as arquivo:
                # Leia o arquivo linha por linha
                linhas = arquivo.readlines()

                for linha in linhas:
                    if "Valor do Empenho: R$" in linha:
                        # print(linha.split("R$")[1].strip() if "R$" in linha else None)
                        valor_empenho = (
                            linha.split("R$")[1].strip() if "R$" in linha else None
                        )
                        VALOR_EMP.insert(0, valor_empenho)
                        break
                if valor_empenho is None:
                    VALOR_EMP.insert(0, "R$ Val_NULO_")

        except StopIteration:
            print("A linha desejada não foi encontrada no arquivo.")
        except Exception as e:
            print(f"Ocorreu um erro Em VALOR DO EMPENHO: {str(e)}")

        ############################################################################################
        try:
            # CAPTURANDO "NOTA/ NUMERO DE EMPENHO" - COLUNA 'D'
            qtd = 0
            time.sleep(1)
            with open(output_file, "r") as arquivo:
                # Leia o arquivo linha por linha
                linhas = arquivo.readlines()
                encontrou_nota_empenho = False
                for linha in linhas:
                    # Verifique se a linha contém "Nota de Empenho"
                    if "Nota de Empenho" in linha:
                        # final_2.insert(1, linha)
                        N_EMPRENHO.append(linha)
                        encontrou_nota_empenho = True
                        break

                # Adiciona o valor padrão se "Nota de Empenho" não for encontrado
                if not encontrou_nota_empenho:
                    #final_2.insert(1, "Nota_Emp: 0000000000000")
                    N_EMPRENHO.insert(0, "Nota_Emp: 0000000000000")

        except Exception as e:
            print(f"Ocorreu um erro ao processar 'Nota de Empenho': {str(e)}")

        try:
            # CAPTANDO "Unidade Orçamentária" COL 'E'
            linha_corrigida = None
            time.sleep(1)
            with open(caminho_arquivo, "r") as arquivo:
                # Leia o arquivo linha por linha
                linhas = arquivo.readlines()
                position = 2
                for linha in linhas:
                    # Verifique se a linha contém "Unidade Orçamentária"
                    if "Unidade Or" in linha:
                        linha_corrigida = linha.replace(
                            "Unidade Orgamentaria", "Unidade Orçamentária"
                        )
                        linha_corrigida = linha_corrigida.replace(
                            "FINANCAS", "FINANÇAS"
                        )
                        linha_corrigida = linha_corrigida.replace(
                            "ADMINISTRAGAO", "ADMINISTRAÇÃO"
                        )
                        linha_corrigida = linha_corrigida.replace(
                            "EDUCAGAO", "EDUCAÇÃO"
                        )
                        linha_corrigida = linha_corrigida.replace("SAUDE", "SAÚDE")
                        linha_corrigida = linha_corrigida.replace(
                            "ASSISTENCIA", "ASSISTÊNCIA"
                        )
                        linha_corrigida = linha_corrigida.replace(
                            "PROMOGAO E CULTURA", "PROMOÇÃO E CULTURA"
                        )
                        # final_2.insert(position, linha_corrigida)
                        ORGAO.append(linha_corrigida)
                        # print(final_2) # ALL RIGHT UNTIL HERE
                        break
            if linha_corrigida is None:
                #final_2.insert(position, "Un_Orçament - Un_Or_Não_Encontrada")
                ORGAO.insert(position, "Un_Orçament - Un_Or_Não_Encontrada")

        except Exception as e:
            print(f"Ocorreu um erro ao processar 'Unidade Orçamentária': {str(e)}")

        except Exception as e:
            print(f"Ocorreu um erro ao processar o PDF {pdf_path}: {str(e)}")

        ############################################################################################

        try:
            # CAPTANDO ELEMENTO DE DESPESAS   -   COL 'F'
            Elemento_De_Despesa = [
                " OUTROS SERVICOS DE TERCEIROS - PESSOA JURIDICA",
                "OUTROS SERVIGOS DE TERCEIROS - PESSOA JURIDICA",
                " OUTROS SERVIGOS DE TERCEIROS - PESSOA FISICA",
                "MATERIAL DE CONSUMO",
                "DIARIAS - PESSOAL CIVIL",
                "OBRIGAGOES PATRONAIS",
                "VENCIMENTO E VANTAGENS FIXAS - PESSOA CIVIL",
                "DESPESAS DE EXERCICIO ANTERIORES",
                "APOSENTADORIA, RESERVA REMUNERADA E REFORMAS",
                "PENSOES",
                "DIARIAS -",
                "PENSÃO ALIMENTICIA",
                "CAMARA VEREADORES",
                "PRINCIPAL DA DIVIDA POR CONTRATO",
                "OBRIGAÇÕES TRIBUTÁRIAS E CONTRIBUTIVAS",
                "PASSAGENS E DESPESAS COM LOCOMOGAO",
                "VENCIMENTOS E VANTAGENS FIXAS - SERVIDORES",
                "CONTRATAGAO POR TEMPO DETERMINADO",
                "MATERIAL, BEM OU SERVIGOS PARA DISTRIBUIÇÃO GRATUITA",
                "SERVIÇOS DE CONSULTORIA",
                "EQUIPAMENTOS E MATERIAL PERMANENTE",
                "ARRENDAMENTO MERCANTIL",
                "EMPRESTIMOS CONSIGNADOS",
                "OUTROS BENEDICIOS PREVIDENCIARIOS",
                "SENTENGAS JUDICIAIS",
                "AQUISIGAO DE IMOVEIS",
                "OUTROS AUXILIOS FINANCEIROS A PESSOA FÍSICA",
                "JUROS SOBRE DIVIDA POR CONTRATO",
                "PREMIAGOES CULTURAIS",
            ]
            # Abra o arquivo para leitura
            time.sleep(1)
            with open("texto_extraido.txt", "r") as arquivo:
                # Leia o arquivo linha por linha
                linhas = arquivo.readlines()
                inserir_em = 3
                for linha in linhas:
                    correspondencia_encontrada = False  # Variável de controle

                    for frase in Elemento_De_Despesa:
                        # Verifique se a frase de Elemento de Despesa está contida na linha
                        if frase in linha:
                            linha_corrigida = (
                                frase.replace("SERVIGOS", "SERVIÇOS")
                                .replace("PREMIAGOES", "PREMIAÇÕES")
                                .replace("AQUISIGAO", "AQUISIÇÃO")
                                .replace("SENTENGAS", "SENTENÇAS")
                                .replace("CONTRATAGAO", "CONTRATAÇÃO")
                                .replace("LOCOMOGAO", "LOCOMOÇÃO")
                                .replace("OBRIGAGOES", "OBRIGAÇÕES")
                            )
                            # final_2.insert(inserir_em, linha_corrigida)
                            OBJETO.append(linha_corrigida)
                            correspondencia_encontrada = (
                                True  # Marque que uma correspondência foi encontrada
                            )
                            # print(final_2)
                            break  # Pare de verificar após encontrar a correspondência
                    if correspondencia_encontrada:
                        break
            if not correspondencia_encontrada:
                # final_2.insert(inserir_em, 'Elem_Desp_Não_Encontrado')
                OBJETO.append("Elem_Desp_Não_Encontrado")

        except StopIteration:
            print("A linha de Elem_De_Despesa não foi encontrada.")
        except Exception as e:
            print(f"Ocorreu um erro ao Captar Elemento De Despesa: {str(e)}")

        ############################################################################################

        try:
            # CAPTANDO "FONTE DE Recurso" COL 'I'  :
            time.sleep(1)
            with open(caminho_arquivo, "r") as arquivo:
                # Leia o arquivo linha por linha
                linhas = arquivo.readlines()
                put_in = 4
                for linha in linhas:
                    # Verifique a linha:"
                    if "Fonte de Recurso:" in linha:
                        linha_corrigida = (
                            linha.replace("Préprios", "Próprios")
                            .replace("ATENGAO BASICA", "ATENÇÃO BÁSICA")
                            .replace("SANITARIA", "SANITÁRIA")
                            .replace("Prdéprios", "Próprios")
                        )
                        # final_2.insert(put_in, linha_corrigida)
                        RECURSO.append(linha_corrigida)
                        break
            if "Fonte de Recurso" not in linha:
                # final_2.insert(put_in, '- FONTE_Rec')
                RECURSO.append("- FONTE_Rec_0")

        except StopIteration:
            print("A linha desejada não foi encontrada no arquivo.")
        except Exception as e:
            print(f"Ocorreu um erro ao Captar FONTE DE RECURSOS: {str(e)}")

        ########################################################################################################

        try:
            # CAPTANDO "CREDOR(A)"          COL 'G'
            time.sleep(1)
            with open(caminho_arquivo, "r") as arquivo:
                # Leia o arquivo linha por linha
                linhas = arquivo.readlines()
                PUT_IN = 5

                for linha in linhas:
                    # Verifique se a linha contém "Valor do Empenho:"
                    if "Credor(A)" in linha:
                        linha_corrigida = (
                            linha.replace("Enderego", "Endereço")
                            .replace("Endereco", "Endereço")
                            .replace("PANIFICAGAO", "PANIFICAÇÃO")
                            .replace("EIREL!", "EIRELI")
                        )
                        # final_2.insert(PUT_IN, linha_corrigida)
                        CREDOR.append(linha_corrigida)
                        break
                else:
                    final_2.insert(PUT_IN, "S/_CREDOR(A)")
                    CREDOR.append("S/_CREDOR(A)")
        except StopIteration:
            print("A linha 'Credor(A)' não foi encontrada no arquivo.")
        except Exception as e:
            print(f"Ocorreu um erro ao processar CREDOR(A): {str(e)}")

        ########################################################################################################

        # linha = ''
        try:
            # CAPTURANDO DATA DO EMPENHO            COL  'J' :
            encontrou_primeira_ocorrencia = (
                False  # Variável para rastrear se já encontramos a primeira ocorrência
            )
            Six_Pos = 6
            time.sleep(1)
            with open(caminho_arquivo, "r") as arquivo:
                match = None
                for linha in arquivo:
                    match = re.search(r"\bEm: \d{2}/\d{2}/\d{4}\b", linha)
                    if match:
                        # final_2.insert(Six_Pos, match.group())
                        DATA_EMP.append(match.group())
                        break
                if not match:
                    # final_2.insert(Six_Pos, ': 00/00/0000 Não Encontrada')
                    DATA_EMP.append(": 00/00/0000 Não Encontrada")

        except StopIteration:
            print("DATA DO DEMPENHO não encontrada!.")
        except Exception as e:
            print(f"Ocorreu um erro Data do Empenho: {str(e)}")
        ###############################################################

        try:
            # CAPTURANDO DATA NOTA FISCAL:
            time.sleep(1)
            def extrair_data_apos_palavra_chave(caminho_arquivo, palavras_chave):
                # global final_2

                with open(caminho_arquivo, "r") as arquivo:
                    texto = arquivo.read()

                    for palavra_chave in palavras_chave:
                        posicao_palavra_chave = texto.find(palavra_chave)
                        if (
                            posicao_palavra_chave != -1
                        ):  # Se a palavra-chave for encontrada no texto
                            texto_apos_palavra_chave = texto[
                                posicao_palavra_chave + len(palavra_chave) :
                            ]
                            padrao_data = re.compile(r"\d{2}/\d{2}/\d{4}")
                            resultado = padrao_data.search(texto_apos_palavra_chave)

                            if resultado:
                                data_encontrada = resultado.group()
                                # print(f"Palavra-chave: {palavra_chave}, Data encontrada: {data_encontrada}")
                                return data_encontrada  # Retorna a primeira data encontrada após a palavra-chave

                return None  # Retorna None se nenhuma data for encontrada após as palavras-chave

            # Recorrentes nas NFs:
            palavras_chave = [
                "NOTA FISCAL",
                "Nome/Razao",
                "DANFE",
                "Eletrénica",
                "Eletrénico",
                "Fletrénica",
            ]

            primeira_data = extrair_data_apos_palavra_chave(output_file, palavras_chave)
            Seven_Pos = 7
            if primeira_data:
                # print(f"A primeira data encontrada após as palavras-chave é: {primeira_data}")
                # final_2.append(primeira_data)
                # final_2.insert(Seven_Pos, primeira_data)
                DATA_NF.append(primeira_data)
            else:
                # print("Nenhuma data encontrada após as palavras-chave.")
                # final_2.insert(Seven_Pos, 'DATA_Da_NOTA_FISCAL_NÃO_ENCONTRADA')
                DATA_NF.append("DATA_Da_NOTA_FISCAL_NÃO_ENCONTRADA")

        except StopIteration:
            print("Data da NF não foi encontrada.")
        except Exception as e:
            print(f"Ocorreu um erro ao procurar Data NF: {str(e)}")

        ###############################################################

        # CAPTURA VALOR DO COMPROVANTE BB COL 'L' :
        def capturar_valor_comprovante(output_file):
            """Esta função só é chamada dentro do Loop que identifica bancos, caos identifique
          BB no Arquivo .txt"""
            palavras_chave = ["VALOR: R$", "VALOR COBRADO", "VALOR TOTAL"]
            valores_encontrados = []
            #VALOR_COMPR = []
            #time.sleep(1)

            with open(output_file, "r", encoding="utf-8") as arquivo:
                linhas = arquivo.readlines()

                for linha in linhas:
                    for palavra_chave in palavras_chave:
                        if linha.startswith(palavra_chave):
                            valor_numerico = linha.split(palavra_chave)[-1].strip()
                            valores_encontrados.append(valor_numerico)

            # Pegar apenas a última ocorrência
            if valores_encontrados:

                valor_ultima_ocorrencia = valores_encontrados[-1]
                VALOR_COMPR.append(valor_ultima_ocorrencia)
                VALOR_COMPR[0] = VALOR_COMPR[0].replace('€', '0')

            # VERIFICA SE UM DOS ELEMENTOS strings DA LISTA É NUMERICO E FICA SÓ COM ELE.
            posicao_verificar = 1  # Posição a ser verificada nas strings
            for elemento in VALOR_COMPR:
                if len(elemento) > posicao_verificar and elemento[posicao_verificar].isdigit():
                    VALOR_COMPR.clear()
                    VALOR_COMPR.append(elemento)

            # Se nenhum valor for encontrado, adicionar marcador
            if not valores_encontrados or valores_encontrados is None:
                VALOR_COMPR.append("Sem_Valor_Comprov")






        ###############################################################

        # DADOS BANCARIOS   -  DEFINE QUAL BANCO:
        try:

            # SE FOR BANCO CAIXA ECONOMICA FEDERAL:
            agencia = None
            conta = None
            valor_comprovante = None
            data_debito = None
            CONTA = set()
            CONTA.clear()

            #time.sleep(1)
            with open(output_file, "r") as arquivo:
                # Ler as linhas do arquivo
                linhas = arquivo.readlines()

                # Verificar se a agência ou conta estão presentes em alguma das linhas
                for linha in linhas:
                    # SE FOR BANCO DO BRASIL:
                    # elif '- BANCO DO BRASIL -' in linha:
                    if "- BANCO DO BRASIL -" in linha or "Banco do Brasil" in linha:
                        Data_do_pagamentoBB = False
                        AgenciaBB = False
                        Conta_BB = False

                        # Val Comprovante para Banco do Brasil
                        capturar_valor_comprovante(output_file)


                        # FUN ENCONTRA DATA DO CMPROVANTE BB
                        caminho_arquivo = 'texto_extraido.txt'
                        #encontrar_data_apos_palavra_chave(caminho_arquivo, palavras_chave)
                        Data_Pag = encontrar_data_apos_palavra_chave(caminho_arquivo, chave)
                        #print(Data_Pag)



                        # print('Loop BB')

                        variavel_ag = "AGENCIA: Não encontrada"
                        variavel_cont = "CONTA: Não encontrada"

                        # Palavras-chave para buscar
                        Ag_CONT = [
                            "AGENCIA:",
                            "CONTA:",
                            "CONTA"
                            "Conta Origem:",
                            "VALOR:",
                            "DEBITO EM:",
                            "VALOR COBRADO",
                            "DATA DO PAGAMENTO",
                            "DATA DA TRANSFERENCIA",
                            "VALOR TOTAL",
                        ]

                        #
                        # Palavras-chave para buscar
                        palavra_chave_cliente = "CLIENTE:"
                        padrao_ag_conta = re.compile(r"CLIENTE: [^\n]*\n(?:[^\n]*\n)?AGENCIA: ([^\s]+) CONTA: (\d{1,}\.\d{3}-\d{1,})")

                        # Abre o arquivo para leitura
                        with open(output_file, "r") as arquivo:
                            encontrou_cliente = False
                            variavel_conta = None

                            # Itera sobre as linhas do arquivo
                            for linha in arquivo:
                                # Verifica se a palavra-chave "CLIENTE:" está presente na linha
                                if palavra_chave_cliente in linha:
                                    encontrou_cliente = True
                                    # Aplica a expressão regular na linha atual e nas próximas duas linhas
                                    match = padrao_ag_conta.search(linha + arquivo.readline() + arquivo.readline())
                                    if match:
                                        variavel_ag = match.group(1).replace("@", "0")
                                        variavel_conta = match.group(2)
                                        #print(f"AGENCIA: {variavel_ag}")
                                        #print(f"CONTA1: {variavel_conta}")
                                        # AG.append(variavel_ag)
                                        # CONTA.append(variavel_conta[0:8])
                                    else:
                                        #print("Padrão não encontrado.")
                                        variavel_ag = 'Agencia_Não_Captada'
                                        variavel_conta = 'Conta_Não_Encontrada'
                                    break  # Interrompe a leitura após encontrar o cliente

                            if not encontrou_cliente:
                                #print("Cliente não encontrado no arquivo.")
                                pass



                            if variavel_ag not in AG:
                                AG.insert(0, variavel_ag)
                                AgenciaBB = True
                                #print('AG_BB ',AG) # Verifiquei se a lista stava duplicada

                            if variavel_cont not in CONTA:
                                #CONTA.insert(0, variavel_cont[0:7])
                                CONTA.add(variavel_conta[0:8])
                                Conta_BB = False
                                #print('ContaBB',CONTA) # Verifiquei se a lista stava duplicada



                            if Data_Pag not in DATA_PG:
                                DATA_PG.insert(0, Data_Pag)
                                Data_do_pagamentoBB = True
                                #print('DT_PG_BB ',DATA_PG)
                            break


                    # BLOCO CAIXA a diante;
                    # if 'GovConta Caixa' in linha
                    elif "GovConta Caixa" in linha or "Conta Origem:" in linha:
                        # print('CAIXA LOOP')
                        # Função para substituir caracteres indesejados
                        def corrigir_formato(texto):
                            return texto.replace("@", "0").replace("S", "$")



                        # Abrir o arquivo em modo de leitura
                        time.sleep(1)
                        with open(output_file, "r") as arquivo:
                            linha_encontrada = False
                            proxima_linha = None
                            Data_do_pagamento = False
                            Agencia = False
                            Conta_en = False
                            # Iterar sobre cada linha do arquivo
                            for linha in arquivo:
                                # Substituir caracteres indesejados
                                linha = corrigir_formato(linha)

                                # Extrair agência e conta da linha 'Conta Origem'
                                # if "PREFEITURA M CHA PRETA" or "Conta Origem:" in linha:
                                #     partes = re.findall(r"\d+", linha)
                                #     agencia = partes[0]
                                #     conta = partes[2].strip()

                                # TRECHO NOVO:
                                if "Conta Origem:" in linha:
                                    partes = re.findall(r"\d+", linha)

                                    # Verificar se há elementos suficientes em 'partes' antes de acessar índices específicos
                                    if len(partes) >= 3:
                                        agencia = partes[0]
                                        #conta = partes[2].strip()
                                        conta = partes[2] + '-' + partes[-1] if len(partes) >= 3 else None
                                    else:
                                        # Caso não haja elementos suficientes, atribuir a 'agencia' e 'conta' o valor da lista 'partes'
                                        agencia = partes[0] if len(partes) >= 1 else None
                                        #conta = partes[2] if len(partes) >= 3 else None
                                        conta = partes[2] + '-' + partes[-1] if len(partes) >= 3 else None

                                elif "PREFEITURA M CHA PRETA" in linha:
                                    linha_encontrada = True


                                elif linha_encontrada:
                                    # Se a linha anterior foi encontrada, pegar a próxima linha
                                    proxima_linha = linha.strip()
                                    partes = re.findall(r"\d+", proxima_linha)

                                    # Verificar se há elementos suficientes em 'partes' antes de acessar índices específicos
                                    if len(partes) >= 3:
                                        agencia = partes[0]
                                        conta = partes[2].strip()



                                # Extrair data da última linha com data e horário
                                padrao_data_hora = re.search(
                                    r"(\d{2}/\d{2}/\d{4} - \d{2}:\d{2}:\d{2})", linha
                                )
                                if padrao_data_hora:
                                    data_debito = padrao_data_hora.group(1).split(" - ")[0]
                                else:
                                    padrao_data = re.search(r"(\d{2}/\d{2}/\d{4})", linha)
                                    if padrao_data:
                                        data_debito = padrao_data.group(1)

                                # FUN IMPORT - EXTR VALOR COMPROVANTE CAIXA:
                                # 'extrair_valor'
                                if valor_comprovante is None:
                                    valor_comprovante = extrair_valor(output_file)

                            if not data_debito:
                                data_debito = "Sem_Data_Do_Debito"
                            if not agencia:
                                agencia = "Agência_não_encontrada"
                            if not conta:
                                conta = "Conta_não_encontrada"
                            if not valor_comprovante:
                                valor_comprovante = "Sem_Comprovante"

                        if data_debito not in DATA_PG:
                            DATA_PG.insert(0, data_debito)
                            Data_do_pagamento = True
                            #print(f'Data de Débito: {data_debito}')
                            #print('DatPG_CX', DATA_PG)

                        if agencia not in AG:
                            # AGENCIA_S.extend(final_2[9])
                            AG.insert(0,agencia)
                            Agencia = True
                            # print(f'Agência: {agencia}')
                            #print('AgenciaCX: ', AG)


                        if conta not in CONTA:
                            CONTA.add(conta)
                            Conta_en = True
                            #CONTA.append(conta)
                            #print('Conta_CX',CONTA)



                        if valor_comprovante not in VALOR_COMPR:
                            # final_2.insert(11, valor_comprovante)
                            VALOR_COMPR.append(valor_comprovante)
                            #print(f'Valor do Comprovante: {valor_comprovante}')
                            #print('ValCompCX',VALOR_COMPR)

                        # print('CAIXA')
                        break






        except StopIteration:
            print("Valores Bancários podem Não ter sido Encontrados.")
        except Exception as e:
            print(f"Erro Em Dados Bancarios: {str(e)}")

        ###############################################################




        ###################################################################




        ###############################################################
        # try:
        # #     # EXPORTANDO LIST
        #     with open('final_2.pkl', 'wb') as arquivo:
        #         pickle.dump(final_2, arquivo)
        # #     pass
        # #
        # #
        # except StopIteration:
        #      print("A linha desejada não foi encontrada no arquivo.")
        # except Exception as e:
        #      print(f"Ocorreu um erro ao processar arquivo PKL: {str(e)}")

        capturados = []
        capturados = [
            N_EMPRENHO,
            ORGAO,
            OBJETO,
            CREDOR,
            VALOR_EMP,
            RECURSO,
            DATA_EMP,
            DATA_NF,
            DATA_PG,
            AG,
            CONTA,
            VALOR_COMPR,
        ]
        # final_4 = [DATA_PG, AG, CONTA, VALOR_COMPR]

        # PREENCHE LISTAS VAZIAS
        #capturados = [sublista if sublista else ["Falha_Na_Captacao"] for sublista in capturados]


        # TESTANDO CAPTAÇÃO TOTAL:
        # print('TESTANDO CAPTAÇÃO TOTAL:')
        # for index, item in enumerate(capturados):
        #    print(index, item)


        ########################################################################################################
        #
        # sys.exit()
        #
        # # IMPORTANDO:
        # with open('final_2.pkl', 'rb') as arquivo:
        #     final_2 = pickle.load(arquivo)

        # for index, item in enumerate(final_2):
        #    print(index, item)

        # LIDA COM AUXENCIA DO Fim Da Lista:
        # for I in final_2:
        #     if len(final_2) < 12:
        #         final_2.append('Valor Não Encontrado')

        # VALOR DO EMPENHO   -   COL  'H'
        # H = final_2[0]#.split("R$")[1].strip() if "R$" in final_2[0] else None
        H = VALOR_EMP[0]
        # print('Elemento Zero', final_2[0])
        # print(H)



        # NUM EMPENHO    COL   'D'
        # D = ''.join(filter(str.isdigit, final_2[1]))
        D = "".join(filter(str.isdigit, N_EMPRENHO[0]))
        # print('Elemento UM', final_2[1])
        # print(D)

        # UNIDADE ORÇAM "ORGAO" COL  'E'
        # E = final_2[2].split('-')[1].strip()
        # E = (lambda data: data.split('-')[1].strip() if len(data.split('-')) > 1 else None)(final_2[2] if len(final_2) > 2 else "")
        # print('DOIS: ',final_2[2])
        # print(E)

        def tratar_string(string):
            partes = string.split("-")
            if len(partes) == 2:
                return partes[1].strip()
            elif len(partes) > 2:
                # return partes[1].strip() + partes[2].strip()
                return partes[1].strip() + " " + partes[2].strip()
            else:
                return (
                    partes[1] if string.endswith("-") else string.split("-")[0].strip()
                )

        E = tratar_string(ORGAO[0])

        # ELEMENTO DE DESPESAS  COL 'F'     Não precisa Tratar. (Bando De palavras)
        # F = final_2[3]
        F = OBJETO[0]
        # F = tratar_string(final_2[3])
        # print(F)

        # FONTE DE RECURSOS COL  'I'
        # I = final_2[4].split('-')[1].strip()
        # I = (lambda data: data.split('-')[1].strip() if len(data.split('-')) > 1 else None)(final_2[4] if len(final_2) > 2 else "")
        # I = (lambda data: data.split('-')[1].strip() if '-' in data else None)(final_2[4] if len(final_2) > 2 else "")

        # (I) Experimental:
        # I = (lambda data: data.split('-')[1].strip() if '-' in data else final_2[4].split('-')[1].strip())(final_2[4] if len(final_2) > 2 else "")
        # print(I)

        #linha = "| Fonte de Recurso: 0010.00. 0005 Recursos Préprios"
        linha = RECURSO[0]

        # Expressão regular para capturar tudo após o último caractere numérico e ponto
        padrao = re.search(r'[^0-9.]+$', linha)

        RECURSO.clear()

        # Verifica se o padrão foi encontrado e obtém a correspondência
        if padrao:
            resultado = padrao.group().strip()
            #print(resultado)
            RECURSO.append(resultado)
        else:
            #print("Nenhum padrão encontrado.")
            RECURSO.append("Sem_Fonte_De_Recursos")
        I = RECURSO[0]

        #print("OK Até Aqui!")
        # CREDOR    COL - G
        # G = final_2[5].split('Endereço')[0].strip().replace('Credor(A):', '').replace('Credor(A);', '').strip()
        # print(G)

        G = (
            CREDOR[0]
            .split("Endereço")[0]
            .strip()
            .replace("Credor(A):", "")
            .replace("Credor(A);", "")
            .strip()
        )

        # DATA DO EMPenho   COL   'J'
        # print(final_2[6])
        # J = final_2[6].split(':')[1].strip()
        # print(J)

        J = DATA_EMP[0].split(":")[1].strip()

        # DATA DA NOTA FISCAL COL 'K'   (Não precisa Tratar)
        # K = final_2[7]
        # print(K)
        K = DATA_NF[0]



        # INDICE 8 'Data Da Transferencia - Comprov Banc'  		COL 'L'   (No treatment needs)
        # L = final_2[8]
        # print(final_2[8])
        if DATA_PG:
            L = DATA_PG[0]
        elif len(DATA_PG) == 0:
            DATA_PG.append('Falha_PG')
            L = DATA_PG[0]
            #print('Data_PG: ',DATA_PG[0])

        # INDICE 9  AG    COL 'M'   (No treatment needs)
        # M = final_2[9]
        if AG:
            M = AG[0]
        elif len(AG) == 0:
            AG.append('Falha_AG')
            M = AG[0]
            #print('Col M AG; ',M)

        # INDICE 10  CONTa     COL 'N'    (No treatment needs)
        # N = final_2[10]

        CONTA = list(CONTA)
        if CONTA:
            N = CONTA[0]
        elif len(CONTA) == 0:
            CONTA.append('Falha_CT')
            N = CONTA[0]
            #print('CONTA: ',N)

        # INDICE 11		'Val Comprovante'		COL   'O'      (No treatment needs)
        O = ''
        Aviso = "Comprovante_Auxente"
        if VALOR_COMPR:
            O = VALOR_COMPR[0]
            #print('Recebeu')

        elif len(VALOR_COMPR) == 0:
            VALOR_COMPR.append(Aviso)
            #print('Val Comprov Col O: ',O)

        elif VALOR_COMPR == None:
            VALOR_COMPR.append('0000')
            #O.append(VALOR_COMPR[0])
        O = VALOR_COMPR[0]
        #print('VAR: ',O)


        import random

        mensagens = ['BINGO!', 'CONCLUIDO!', 'TERMINAMOS!', 'MAIS UM PRA CONTA😃', ' UM A MENOS HOJE UFA!',
                     'UFA!',"ESTAMOS INDO BEM😃","AS VEZES EU IMPRESSIONO A MIM MESMO."]
        mensagem_aleatoria = random.choice(mensagens)
        print(mensagem_aleatoria)



        final_3 = [D, E, F, I, H, G, J, K, L, M, N, O]



        # PREENCHE ITENS FALTANTES / EVITA LISTAS VAZIAS:
        #final_3 = [sublista if sublista else ["Falha_Na_Captacao"] for sublista in final_3]


        #final_3 = [set(item) for item in final_3]

        # preenchimento = "Falha_Na_Captação"
        # final_3 = (
        #     final_3 + [preenchimento] * (12 - len(final_3))
        #     if len(final_3) < 12
        #     else final_3
        # )

        # for I in final_3:
        #   print(I)

        # for index, item in enumerate(final_3):
        # print(index, item)

        colunas = ["D", "E", "F", "I", "H", "G", "J", "K", "L", "M", "N", "O"]

        # print('Formato Data NF: ',final_2[7])
        # print('Data da NF:', K)
        time.sleep(1)
        # Itere sobre as colunas e os valores
        for coluna, valor in zip(colunas, final_3):
            if valor is not None:
                linha_vazia = 2
                celula_destino = f"{coluna}{linha_vazia}"

                # Encontre a primeira linha vazia na coluna especificada
                while sheet[celula_destino].value is not None:
                    linha_vazia += 1
                    celula_destino = f"{coluna}{linha_vazia}"

                # Preencha a célula com o valor
                sheet[celula_destino] = valor

        # RENOMEIA PDF P/ NUM EMPENHO - (MUDAR CAMINHO PC):
        import random
        import string

        caminho_pasta_pdf = "/home/dikson/PycharmProjects/Vision_Modularizado/PRODUTO_FINAL_13/Renamed"
        print(f"Dados do arquivo {pdf_path} adicionados à planilha.")

        if N_EMPRENHO:
            # Extrair o número da Nota de Empenho para Renomear o Arquivo .PDF
            numero_empenho = "".join(filter(str.isdigit, N_EMPRENHO[0]))

            if N_EMPRENHO[0] == "Nota_Emp: 0000000000000":
                sufixo_aleatorio = "".join(
                    random.choices(string.ascii_letters + string.digits, k=6)
                )
                novo_nome_arquivo = f"{numero_empenho}_{sufixo_aleatorio}.PDF"
            else:
                novo_nome_arquivo = f"{numero_empenho}.PDF"

            caminho_novo_arquivo = os.path.join(caminho_pasta_pdf, novo_nome_arquivo)
            os.rename(pdf_path, caminho_novo_arquivo)  # Renomeie o arquivo

            print(f"Arquivo {pdf_path} renomeado para {novo_nome_arquivo}")

        # LIMPAR LISTAS IDIVIDUALMENTE:
        time.sleep(1)

        ORGAO.clear()
        OBJETO.clear()
        CREDOR.clear()
        VALOR_EMP.clear()
        RECURSO.clear()
        DATA_EMP.clear()
        DATA_NF.clear()
        DATA_PG.clear()
        AG.clear()
        CONTA.clear()
        VALOR_COMPR.clear()
        N_EMPRENHO.clear()

        # Limpar a lista de listas (COLETIVAMENTE)
        # for item in final_3:
        #     item.clear()


    except StopIteration:
        print(
            "Não foi possivel identificar algumas informações do PDF', Algumas celulas ficarão vazias :/ O programa continuará."
        )
    except Exception as e:
        print(f"Ocorreu um erro ao processar o PDF {pdf_path}: {str(e)}")


# Função para selecionar arquivos PDF com uma janela de diálogo
def selecionar_arquivos_pdf():
    root = tk.Tk()
    root.withdraw()
    file_paths = filedialog.askopenfilenames(
        title="Selecione arquivos PDF", filetypes=[("PDF Files", "*.PDF")]
    )

    if file_paths:
        workbook = openpyxl.load_workbook("despesa.xlsx")

        # Selecione a folha em que deseja adicionar os dados (por exemplo, a primeira folha)
        sheet = workbook.active

        final_2 = []

        # Itere sobre os arquivos PDF selecionados e processe cada um
        for pdf_path in file_paths:
            processar_pdf(pdf_path, sheet)
            # final_2.extend(result)

        # Salve as alterações no arquivo
        workbook.save("despesa.xlsx")

        # Feche o arquivo
        workbook.close()


# Chame a função para selecionar arquivos PDF
selecionar_arquivos_pdf()

tempo_final = time.time()
tempo_decorrido = tempo_final - tempo_inicial
Hora_final = datetime.datetime.now()
Hora_final = Hora_final.strftime("%H:%M")

messagebox.showinfo(
    f"Processo(s) Finalizado(s) !\n",
    f"Tempo levado: {tempo_decorrido:.2f} segundos \n"
    f" Iniciado em ;{Hora_inicial} \n"
    f" Finalizado em; {Hora_final}",
)
print('TODAS AS TAREFAS CONCLUIDAS 🌒')

